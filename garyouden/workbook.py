from __future__ import annotations

from collections.abc import Iterable, Iterator
from pathlib import Path
from typing import Any
from unicodedata import category

from garyouden.layout import ARMY_UNIT_ROLES
from garyouden.models import DatFile, Save
from garyouden.summaries import city_name, monarch_name, none_if_free_power_id, power_ruler_name


ARMY_UNIT_LABELS = {
    "commander": "主将",
    "vanguard": "先锋",
    "left_wing": "左翼",
    "right_wing": "右翼",
    "left_reserve": "左备",
    "right_reserve": "右备",
}


def write_dat_workbook(
    dat_file: DatFile,
    path: str | Path,
    saves: Iterable[Save] | None = None,
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    selected = dat_file.saves if saves is None else tuple(saves)
    workbook = Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    first_sheet = workbook.active
    first_sheet.title = "存档信息"
    write_table(first_sheet, save_headers(), save_rows(dat_file, selected), header_font, header_fill)

    sheet_specs = (
        ("势力", power_headers(), power_rows(selected)),
        ("外交", diplomacy_headers(), diplomacy_rows(selected)),
        ("城池信息", city_headers(), city_rows(selected)),
        ("军团信息", army_headers(), army_rows(selected)),
        ("君主信息", monarch_headers(), monarch_rows(selected)),
        ("未知区块", unknown_block_headers(), unknown_block_rows(selected)),
    )
    for title, headers, rows in sheet_specs:
        worksheet = workbook.create_sheet(title)
        write_table(worksheet, headers, rows, header_font, header_fill)

    workbook_path = Path(path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)
    return workbook_path


def write_table(
    worksheet: Any,
    headers: list[str],
    rows: Iterable[tuple[Any, ...]],
    header_font: Any,
    fill: Any,
) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append(tuple(excel_safe_value(value) for value in row))

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)


def autosize_columns(worksheet: Any) -> None:
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)


def excel_safe_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    return "".join(escape_control_char(char) for char in value)


def escape_control_char(char: str) -> str:
    codepoint = ord(char)
    if char in "\t\n\r":
        return char
    if codepoint < 0x20 or category(char) == "Cc":
        return f"\\x{codepoint:02X}"
    return char


def save_headers() -> list[str]:
    return [
        "存档路径",
        "存档序号",
        "剧本名/存档名",
        "日期",
        "起始年",
        "起始月",
        "起始日",
        "编号",
        "玩家势力",
        "玩家势力君主",
        "信赖度",
        "本月税率",
        "本月征兵1",
        "本月征兵2",
        "本月征兵3",
        "下月税率",
        "下月征兵1",
        "下月征兵2",
        "下月征兵3",
        "总势力数",
        "非空势力槽",
        "已知城池",
        "非空军团槽",
        "登场君主",
    ]


def save_rows(dat_file: DatFile, saves: Iterable[Save]) -> Iterator[tuple[Any, ...]]:
    for save in saves:
        header = save.header
        player_power_id = none_if_free_power_id(header.player_power_id)
        yield (
            str(dat_file.path),
            save.index,
            header.name,
            f"{header.year:04d}-{header.month:02d}-{header.day:02d}",
            header.year,
            header.month,
            header.day,
            header.save_id,
            player_power_id,
            power_ruler_name(save, header.player_power_id),
            header.trust,
            header.this_month_tax,
            *header.this_month_recruits,
            header.next_month_tax,
            *header.next_month_recruits,
            header.power_count,
            len(save.active_powers),
            len(save.known_cities),
            len(save.active_armies),
            len(save.appeared_monarchs),
        )


def power_headers() -> list[str]:
    return [
        "存档序号",
        "势力序号",
        "状态",
        "是否有效",
        "君主编号",
        "君主姓名",
        "军师编号",
        "军师姓名",
        "首都编号",
        "首都名称",
        "起始兵数1",
        "起始兵数2",
        "起始兵数3",
        "武将数量",
        "钱",
        "城池数",
        "外交官编号",
        "外交官姓名",
    ]


def power_rows(saves: Iterable[Save]) -> Iterator[tuple[Any, ...]]:
    for save in saves:
        for power in save.powers:
            yield (
                save.index,
                power.index,
                power.status,
                power.is_active,
                power.ruler_monarch_id,
                monarch_name(save, power.ruler_monarch_id),
                power.strategist_monarch_id,
                monarch_name(save, power.strategist_monarch_id),
                power.capital_city_id,
                city_name(save, power.capital_city_id),
                *power.starting_troops,
                power.monarch_count,
                power.money,
                power.city_count,
                power.diplomat_monarch_id,
                monarch_name(save, power.diplomat_monarch_id),
            )


def diplomacy_headers() -> list[str]:
    return ["存档序号", "势力序号", "势力君主", "目标势力序号", "目标势力君主", "友好度"]


def diplomacy_rows(saves: Iterable[Save]) -> Iterator[tuple[Any, ...]]:
    for save in saves:
        for source_power_id, relations in enumerate(save.diplomacy):
            for target_power_id, friendship in enumerate(relations):
                yield (
                    save.index,
                    source_power_id,
                    power_ruler_name(save, source_power_id),
                    target_power_id,
                    power_ruler_name(save, target_power_id),
                    friendship,
                )


def city_headers() -> list[str]:
    return [
        "存档序号",
        "城池序号",
        "状态",
        "所属势力序号",
        "所属势力君主",
        "城池名称",
        "坐标原始值",
        "最大生产力",
        "当前生产力",
        "上升值",
        "防灾值",
        "城兵",
        "类型值",
        "类型",
        "内政官序号",
        "内政官姓名",
    ]


def city_rows(saves: Iterable[Save]) -> Iterator[tuple[Any, ...]]:
    for save in saves:
        for city in save.cities:
            yield (
                save.index,
                city.index,
                city.status,
                none_if_free_power_id(city.owner_power_id),
                power_ruler_name(save, city.owner_power_id),
                city.name,
                bytes_to_hex(city.coordinate_bytes),
                city.max_productivity,
                city.current_productivity,
                city.growth_value,
                city.disaster_resistance,
                city.garrison_troops,
                city.city_type_value,
                city.city_type_label,
                city.civil_monarch_id,
                monarch_name(save, city.civil_monarch_id),
            )


def army_headers() -> list[str]:
    headers = [
        "存档序号",
        "军团序号",
        "状态",
        "是否有效",
        "所属势力序号",
        "所属势力君主",
        "军团长序号",
        "军团长姓名",
        "总兵数",
        "士气",
        "当前坐标原始值",
        "目标坐标原始值",
        "目的地城池序号",
        "目的地城池名称",
    ]
    for role in ARMY_UNIT_ROLES:
        label = ARMY_UNIT_LABELS[role]
        headers.extend([f"{label}兵数", f"{label}兵种", f"{label}原始值"])
    return headers


def army_rows(saves: Iterable[Save]) -> Iterator[tuple[Any, ...]]:
    for save in saves:
        for army in save.armies:
            row = [
                save.index,
                army.index,
                army.status,
                army.is_active,
                none_if_free_power_id(army.owner_power_id),
                power_ruler_name(save, army.owner_power_id),
                army.commander_monarch_id,
                monarch_name(save, army.commander_monarch_id),
                army.total_troops,
                army.morale,
                bytes_to_hex(army.current_coordinate_bytes),
                bytes_to_hex(army.target_coordinate_bytes),
                army.destination_city_id,
                city_name(save, army.destination_city_id),
            ]
            for unit in army.units:
                row.extend([unit.troops, unit.troop_type, bytes_to_hex(unit.raw)])
            yield tuple(row)


def monarch_headers() -> list[str]:
    return [
        "存档序号",
        "君主序号",
        "属性",
        "不会被俘",
        "是否君主",
        "是否登场",
        "头像序号",
        "姓名",
        "号",
        "城塞战",
        "野战",
        "水战",
        "武力",
        "统率",
        "政治",
        "列表状态值",
        "列表状态",
        "登场倒计时",
        "投奔势力",
        "投奔势力君主",
        "所属势力",
        "所属势力君主",
        "被俘所属势力",
        "被俘所属势力君主",
    ]


def monarch_rows(saves: Iterable[Save]) -> Iterator[tuple[Any, ...]]:
    for save in saves:
        for monarch in save.monarchs:
            yield (
                save.index,
                monarch.index,
                monarch.attributes,
                monarch.avoids_capture,
                monarch.is_ruler,
                monarch.is_appeared,
                monarch.portrait_id,
                monarch.name,
                monarch.courtesy_name,
                monarch.siege_ability,
                monarch.field_ability,
                monarch.naval_ability,
                monarch.prowess,
                monarch.command,
                monarch.politics,
                monarch.list_status_value,
                monarch.list_status_label,
                monarch.appearance_countdown,
                none_if_free_power_id(monarch.target_power_id),
                power_ruler_name(save, monarch.target_power_id),
                none_if_free_power_id(monarch.owner_power_id),
                power_ruler_name(save, monarch.owner_power_id),
                none_if_free_power_id(monarch.captive_owner_power_id),
                power_ruler_name(save, monarch.captive_owner_power_id),
            )


def unknown_block_headers() -> list[str]:
    return ["存档序号", "大项", "偏移", "大小", "非零字节数"]


def unknown_block_rows(saves: Iterable[Save]) -> Iterator[tuple[Any, ...]]:
    for save in saves:
        yield (
            save.index,
            "军团前未知区块",
            f"0x{save.pre_army_block.offset:04X}",
            save.pre_army_block.size,
            save.pre_army_block.nonzero_bytes,
        )
        yield (
            save.index,
            "尾部未知区块",
            f"0x{save.tail_block.offset:04X}",
            save.tail_block.size,
            save.tail_block.nonzero_bytes,
        )


def bytes_to_hex(values: bytes | tuple[int, ...]) -> str:
    return " ".join(f"{value:02X}" for value in values)
